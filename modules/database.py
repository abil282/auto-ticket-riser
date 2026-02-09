"""
Database and Ticket Storage Module
Manages ticket storage in SQLite database and Excel files
"""

import sqlite3
import json
import os
from datetime import datetime
import uuid
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class TicketDatabase:
    """Manage ticket storage in SQLite database"""
    
    def __init__(self, db_path='data/tickets/tickets.db'):
        """Initialize database connection"""
        self.db_path = db_path
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        self.conn = None
        self.initialize_database()
    
    def initialize_database(self):
        """Create necessary tables if they don't exist"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Tickets table (include asset_id for system/asset tracking)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS tickets (
                ticket_id TEXT PRIMARY KEY,
                user_name TEXT NOT NULL,
                user_email TEXT NOT NULL,
                department TEXT,
                phone TEXT,
                asset_id TEXT,
                original_description TEXT NOT NULL,
                corrected_description TEXT NOT NULL,
                category TEXT NOT NULL,
                priority TEXT NOT NULL,
                status TEXT DEFAULT 'Open',
                assigned_to TEXT,
                created_timestamp TEXT NOT NULL,
                updated_timestamp TEXT NOT NULL,
                resolved_timestamp TEXT,
                resolution_notes TEXT,
                attachments TEXT,
                metadata TEXT
            )
        ''')
        
        # Ticket history/audit table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ticket_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id TEXT NOT NULL,
                action TEXT NOT NULL,
                performed_by TEXT NOT NULL,
                timestamp TEXT NOT NULL,
                details TEXT,
                FOREIGN KEY(ticket_id) REFERENCES tickets(ticket_id)
            )
        ''')
        
        # Statistics table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS statistics (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                total_tickets INTEGER,
                resolved_tickets INTEGER,
                open_tickets INTEGER,
                high_priority INTEGER,
                category TEXT
            )
        ''')
        
        # Ensure legacy DBs get asset_id column
        conn.commit()
        cursor.execute("PRAGMA table_info(tickets)")
        cols = [r[1] for r in cursor.fetchall()]
        if 'asset_id' not in cols:
            try:
                cursor.execute('ALTER TABLE tickets ADD COLUMN asset_id TEXT')
            except Exception:
                pass
        conn.commit()
        conn.close()
    
    def create_ticket(self, ticket_data):
        """
        Create a new ticket
        
        Args:
            ticket_data (dict): Ticket information
            
        Returns:
            str: Ticket ID
        """
        ticket_id = f"TKT-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        now = datetime.now().isoformat()
        
        cursor.execute('''
            INSERT INTO tickets (
                ticket_id, user_name, user_email, department, phone, asset_id,
                original_description, corrected_description,
                category, priority, status, assigned_to,
                created_timestamp, updated_timestamp, metadata
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            ticket_id,
            ticket_data.get('user_name', ''),
            ticket_data.get('user_email', ''),
            ticket_data.get('department', ''),
            ticket_data.get('phone', ''),
            ticket_data.get('asset_id', ''),
            ticket_data.get('original_description', ''),
            ticket_data.get('corrected_description', ''),
            ticket_data.get('category', 'General'),
            ticket_data.get('priority', 'P3 - Medium'),
            'Open',
            ticket_data.get('assigned_to', 'Unassigned'),
            now,
            now,
            json.dumps(ticket_data.get('metadata', {}))
        ))
        
        conn.commit()
        self.add_history(ticket_id, 'Created', 'System', 'Ticket created', conn)
        conn.close()
        
        return ticket_id
    
    def get_ticket(self, ticket_id):
        """Get ticket by ID"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM tickets WHERE ticket_id = ?', (ticket_id,))
        row = cursor.fetchone()
        conn.close()
        
        if row:
            return dict(row)
        return None
    
    def get_all_tickets(self, filters=None):
        """Get all tickets with optional filters"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        query = 'SELECT * FROM tickets WHERE 1=1'
        params = []
        
        if filters:
            if filters.get('status'):
                query += ' AND status = ?'
                params.append(filters['status'])
            if filters.get('category'):
                query += ' AND category = ?'
                params.append(filters['category'])
            if filters.get('priority'):
                query += ' AND priority = ?'
                params.append(filters['priority'])
            if filters.get('assigned_to'):
                query += ' AND assigned_to = ?'
                params.append(filters['assigned_to'])
            if filters.get('date_from'):
                query += ' AND DATE(created_timestamp) >= ?'
                params.append(filters['date_from'])
            if filters.get('date_to'):
                query += ' AND DATE(created_timestamp) <= ?'
                params.append(filters['date_to'])
        
        query += ' ORDER BY created_timestamp DESC'
        
        cursor.execute(query, params)
        tickets = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        return tickets
    
    def update_ticket(self, ticket_id, updates):
        """Update ticket details"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        performed_by = updates.pop('performed_by', 'System')
        updates['updated_timestamp'] = datetime.now().isoformat()

        set_clause = ', '.join([f'{key} = ?' for key in updates.keys()])
        values = list(updates.values()) + [ticket_id]

        cursor.execute(f'UPDATE tickets SET {set_clause} WHERE ticket_id = ?', values)
        conn.commit()

        self.add_history(ticket_id, 'Updated', performed_by, f'Updated: {", ".join(updates.keys())}', conn)
        conn.close()
    
    def add_history(self, ticket_id, action, performed_by, details, conn=None):
        """Add ticket to history"""
        close_conn = False
        if conn is None:
            conn = sqlite3.connect(self.db_path)
            close_conn = True
        
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO ticket_history (ticket_id, action, performed_by, timestamp, details)
            VALUES (?, ?, ?, ?, ?)
        ''', (ticket_id, action, performed_by, datetime.now().isoformat(), details))
        
        conn.commit()
        
        if close_conn:
            conn.close()
    
    def get_ticket_history(self, ticket_id):
        """Get ticket history"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT * FROM ticket_history WHERE ticket_id = ?
            ORDER BY timestamp DESC
        ''', (ticket_id,))
        
        history = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        return history


class ExcelReportGenerator:
    """Generate Excel reports from ticket data"""
    
    def __init__(self, output_dir='data/reports'):
        """Initialize report generator"""
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def generate_ticket_report(self, tickets, report_type='all', filename=None):
        """
        Generate Excel report with ticket data
        
        Args:
            tickets (list): List of ticket dictionaries
            report_type (str): 'all', 'category', or 'date'
            filename (str): Custom filename
            
        Returns:
            str: Path to generated file
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tickets"
        
        # Define headers
        headers = [
            'Ticket ID', 'User Name', 'Email', 'Department',
            'Original Issue', 'Corrected Issue', 'Category', 'Priority',
            'Status', 'Assigned To', 'Created Date', 'Updated Date'
        ]
        
        # Add headers with styling
        self._add_header_row(ws, headers)
        
        # Add data
        for idx, ticket in enumerate(tickets, start=2):
            ws[f'A{idx}'] = ticket.get('ticket_id', '')
            ws[f'B{idx}'] = ticket.get('user_name', '')
            ws[f'C{idx}'] = ticket.get('user_email', '')
            ws[f'D{idx}'] = ticket.get('department', '')
            ws[f'E{idx}'] = ticket.get('original_description', '')[:100]  # Truncate
            ws[f'F{idx}'] = ticket.get('corrected_description', '')[:100]
            ws[f'G{idx}'] = ticket.get('category', '')
            ws[f'H{idx}'] = ticket.get('priority', '')
            ws[f'I{idx}'] = ticket.get('status', '')
            ws[f'J{idx}'] = ticket.get('assigned_to', '')
            ws[f'K{idx}'] = ticket.get('created_timestamp', '')[:10]
            ws[f'L{idx}'] = ticket.get('updated_timestamp', '')[:10]
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Generate filename
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"tickets_report_{report_type}_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    def generate_category_report(self, tickets):
        """Generate category-wise report"""
        # Group by category
        by_category = {}
        for ticket in tickets:
            cat = ticket.get('category', 'General')
            if cat not in by_category:
                by_category[cat] = []
            by_category[cat].append(ticket)
        
        wb = openpyxl.Workbook()
        
        for category, cat_tickets in by_category.items():
            ws = wb.create_sheet(title=category[:31])  # Sheet name limit
            
            headers = [
                'Ticket ID', 'User', 'Issue', 'Priority', 'Status', 'Created'
            ]
            self._add_header_row(ws, headers)
            
            for idx, ticket in enumerate(cat_tickets, start=2):
                ws[f'A{idx}'] = ticket.get('ticket_id', '')
                ws[f'B{idx}'] = ticket.get('user_name', '')
                ws[f'C{idx}'] = ticket.get('corrected_description', '')[:80]
                ws[f'D{idx}'] = ticket.get('priority', '')
                ws[f'E{idx}'] = ticket.get('status', '')
                ws[f'F{idx}'] = ticket.get('created_timestamp', '')[:10]
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filepath = os.path.join(self.output_dir, f"tickets_by_category_{timestamp}.xlsx")
        wb.save(filepath)
        
        return filepath
    
    def generate_date_wise_report(self, tickets):
        """Generate date-wise report"""
        # Group by date
        by_date = {}
        for ticket in tickets:
            date = ticket.get('created_timestamp', 'Unknown')[:10]
            if date not in by_date:
                by_date[date] = []
            by_date[date].append(ticket)
        
        wb = openpyxl.Workbook()
        
        for date, date_tickets in sorted(by_date.items()):
            ws = wb.create_sheet(title=date)
            
            headers = [
                'Ticket ID', 'User', 'Category', 'Priority', 'Status', 'Issue'
            ]
            self._add_header_row(ws, headers)
            
            for idx, ticket in enumerate(date_tickets, start=2):
                ws[f'A{idx}'] = ticket.get('ticket_id', '')
                ws[f'B{idx}'] = ticket.get('user_name', '')
                ws[f'C{idx}'] = ticket.get('category', '')
                ws[f'D{idx}'] = ticket.get('priority', '')
                ws[f'E{idx}'] = ticket.get('status', '')
                ws[f'F{idx}'] = ticket.get('corrected_description', '')[:80]
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filepath = os.path.join(self.output_dir, f"tickets_by_date_{timestamp}.xlsx")
        wb.save(filepath)
        
        return filepath
    
    def _add_header_row(self, ws, headers):
        """Add styled header row"""
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
    
    def generate_summary_report(self, tickets):
        """Generate summary report with statistics"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        
        # Statistics
        total_tickets = len(tickets)
        open_tickets = len([t for t in tickets if t.get('status') == 'Open'])
        assigned_tickets = len([t for t in tickets if t.get('status') == 'Assigned'])
        in_progress = len([t for t in tickets if t.get('status') == 'In Progress'])
        resolved = len([t for t in tickets if t.get('status') == 'Resolved'])
        closed = len([t for t in tickets if t.get('status') == 'Closed'])
        
        # Category breakdown
        categories = {}
        for ticket in tickets:
            cat = ticket.get('category', 'General')
            categories[cat] = categories.get(cat, 0) + 1
        
        # Priority breakdown
        priorities = {}
        for ticket in tickets:
            pri = ticket.get('priority', 'P3 - Medium')
            priorities[pri] = priorities.get(pri, 0) + 1
        
        # Summary Section
        row = 1
        ws[f'A{row}'] = "TICKET SUMMARY REPORT"
        ws[f'A{row}'].font = Font(bold=True, size=14, color='FFFFFF')
        ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 2
        ws[f'A{row}'] = "Status Summary"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        
        row += 1
        ws[f'A{row}'] = "Total Tickets"
        ws[f'B{row}'] = total_tickets
        
        row += 1
        ws[f'A{row}'] = "Open"
        ws[f'B{row}'] = open_tickets
        
        row += 1
        ws[f'A{row}'] = "Assigned"
        ws[f'B{row}'] = assigned_tickets
        
        row += 1
        ws[f'A{row}'] = "In Progress"
        ws[f'B{row}'] = in_progress
        
        row += 1
        ws[f'A{row}'] = "Resolved"
        ws[f'B{row}'] = resolved
        
        row += 1
        ws[f'A{row}'] = "Closed"
        ws[f'B{row}'] = closed
        
        # Category breakdown
        row += 2
        ws[f'A{row}'] = "Category Breakdown"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        
        row += 1
        for category, count in sorted(categories.items()):
            ws[f'A{row}'] = category
            ws[f'B{row}'] = count
            row += 1
        
        # Priority breakdown
        row += 2
        ws[f'A{row}'] = "Priority Breakdown"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        
        row += 1
        for priority, count in sorted(priorities.items()):
            ws[f'A{row}'] = priority
            ws[f'B{row}'] = count
            row += 1
        
        # Adjust columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filepath = os.path.join(self.output_dir, f"tickets_summary_{timestamp}.xlsx")
        wb.save(filepath)
        
        return filepath
    
    def generate_history_report(self, tickets):
        """Generate report with ticket history and audit trail"""
        db = TicketDatabase()
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Summary", 0)
        summary_ws['A1'] = "TICKET HISTORY REPORT"
        summary_ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        summary_ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        summary_ws.merge_cells('A1:H1')
        
        summary_ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        summary_ws['A3'] = f"Total Tickets: {len(tickets)}"
        
        row = 5
        headers = ['Ticket ID', 'User', 'Category', 'Status', 'Priority', 'Assigned To', 'Created', 'Updated']
        for col_num, header in enumerate(headers, 1):
            cell = summary_ws.cell(row=row, column=col_num)
            cell.value = header
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.font = Font(bold=True)
        
        row = 6
        for ticket in tickets:
            summary_ws[f'A{row}'] = ticket.get('ticket_id', '')
            summary_ws[f'B{row}'] = ticket.get('user_name', '')
            summary_ws[f'C{row}'] = ticket.get('category', '')
            summary_ws[f'D{row}'] = ticket.get('status', '')
            summary_ws[f'E{row}'] = ticket.get('priority', '')
            summary_ws[f'F{row}'] = ticket.get('assigned_to', '')
            summary_ws[f'G{row}'] = str(ticket.get('created_timestamp', ''))[:10]
            summary_ws[f'H{row}'] = str(ticket.get('updated_timestamp', ''))[:10]
            row += 1
        
        # Create detail sheets for each ticket with history
        for ticket in tickets[:20]:  # Limit to first 20 to avoid huge files
            ticket_id = ticket.get('ticket_id', 'Unknown')
            sheet_name = ticket_id[-10:] if len(ticket_id) > 31 else ticket_id
            
            ws = wb.create_sheet(title=sheet_name)
            
            # Ticket details
            ws['A1'] = f"Ticket: {ticket_id}"
            ws['A1'].font = Font(bold=True, size=12)
            
            row = 3
            ws[f'A{row}'] = "User"
            ws[f'B{row}'] = ticket.get('user_name', '')
            
            row += 1
            ws[f'A{row}'] = "Email"
            ws[f'B{row}'] = ticket.get('user_email', '')
            
            row += 1
            ws[f'A{row}'] = "Category"
            ws[f'B{row}'] = ticket.get('category', '')
            
            row += 1
            ws[f'A{row}'] = "Priority"
            ws[f'B{row}'] = ticket.get('priority', '')
            
            row += 1
            ws[f'A{row}'] = "Status"
            ws[f'B{row}'] = ticket.get('status', '')
            
            row += 1
            ws[f'A{row}'] = "Assigned To"
            ws[f'B{row}'] = ticket.get('assigned_to', '')
            
            row += 1
            ws[f'A{row}'] = "Created"
            ws[f'B{row}'] = ticket.get('created_timestamp', '')
            
            row += 2
            ws[f'A{row}'] = "ISSUE DESCRIPTION"
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = ticket.get('corrected_description', '')
            
            row += 2
            ws[f'A{row}'] = "ACTIVITY HISTORY"
            ws[f'A{row}'].font = Font(bold=True)
            
            # Get history
            history = db.get_ticket_history(ticket_id)
            row += 1
            
            if history:
                ws[f'A{row}'] = "Date"
                ws[f'B{row}'] = "Action"
                ws[f'C{row}'] = "By"
                row += 1
                
                for entry in history:
                    ws[f'A{row}'] = entry.get('timestamp', '')
                    ws[f'B{row}'] = entry.get('action', '')
                    ws[f'C{row}'] = entry.get('performed_by', '')
                    row += 1
            else:
                ws[f'A{row}'] = "No history available"
            
            # Adjust columns
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 20
        
        # Adjust summary sheet columns
        for col in summary_ws.columns:
            summary_ws.column_dimensions[get_column_letter(col[0].column)].width = 18
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filepath = os.path.join(self.output_dir, f"tickets_with_history_{timestamp}.xlsx")
        wb.save(filepath)
        
        return filepath

